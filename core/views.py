import datetime
import io
import os
import shutil

import xlsxwriter
from django.http import HttpResponse
from django_filters import rest_framework as filters
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.decorators import api_view
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from api.settings import TEMPLATE_ROOT_VOTE_FILE
from api.settings import UPLOAD_DIR_ELECTOR, MEDIA_URL
from globals.pagination import CustomPagination
from globals.utils import format_form_errors_response
from users.constants import ROLE_SUPERVISOR
from .filters import CampaignFilter, ElectorFilter, VoteFilter, VoteOfficeFilter
from .forms import UploadFileForm, HaveVoteForm, AddVoteOffice
from .helpers import handle_uploaded_file, compare_pictures, decode_base64_file, get_elector_pictures, send_mt
from .models import Elector, Vote, Campaign, VoteOffice
from .serializers import ElectorSerializers, VoteSerializers, CampaignSerializers, UploadFileSerializer, \
    VoteOfficeSerializers

date_style = NamedStyle(name='date', number_format='DD/MM/YYYY')
time_style = NamedStyle(name='time', number_format='HH:MM:SS')
datetime_style = NamedStyle(name='time', number_format='DD/MM/YYYY HH:MM:SS')


# Create your views here.


@api_view(['POST'])
def upload_elector_picture_view(request):
    if not request.FILES:
        form = UploadFileSerializer(data=request.data)
    else:
        form = UploadFileForm(request.data, request.FILES)
    if form.is_valid():
        file_dir = handle_uploaded_file(
            request.FILES['picture'] if request.FILES else decode_base64_file(request.data['picture']))
        response = {'match': False}
        if os.path.exists(file_dir):
            directory, filename, matricule = compare_pictures(file_dir)
            if matricule:
                response['match'] = True
                response['url'] = MEDIA_URL + directory + '/' + filename
                serializer = ElectorSerializers(Elector.objects.filter(matricule__iexact=matricule).last(), many=False)
                response['elector'] = serializer.data

        return Response(status=status.HTTP_200_OK, data=response)
    return Response(
        data=format_form_errors_response(form.errors),
        status=status.HTTP_400_BAD_REQUEST
    )


@api_view(['GET'])
def get_elector_view(request):
    elector = Elector.objects.filter(matricule=request.GET.get('matricule', None)).last()
    response = {'match': False}
    if not elector:
        return Response(status=status.HTTP_200_OK, data=response)
    matricule = elector.matricule
    pictures = get_elector_pictures(elector.matricule)
    response['match'] = True
    if pictures:
        response['url'] = MEDIA_URL + matricule + '/' + pictures[0]
    user = request.user
    if user.role == ROLE_SUPERVISOR:
        voteOffice = user.voteOffice
        if not voteOffice:
            elector = None
        else:
            elector = Elector.objects.filter(deleted=False, voteOffice__id=voteOffice.id,
                                             matricule__iexact=matricule).last()
    else:
        elector = Elector.objects.filter(matricule__iexact=matricule).last()
    if not elector:
        response['match'] = False
        del response['url']
    else:
        serializer = ElectorSerializers(elector, many=False)
        response['elector'] = serializer.data
    return Response(status=status.HTTP_200_OK, data=response)


class VoteView(viewsets.ReadOnlyModelViewSet):
    queryset = Vote.objects.all()
    serializer_class = VoteSerializers
    pagination_class = CustomPagination
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = VoteFilter
    permission_classes = (IsAuthenticated,)

    @action(detail=False, methods=['post'], url_path='has-vote')
    def has_vote(self, request, *args, **kwargs):
        form = HaveVoteForm(data=request.data or {})
        if form.is_valid():
            have_vote = form.cleaned_data
            vote = Vote.objects.filter(voted=True, campaign__id=have_vote['campaign'].id,
                                       elector__id=have_vote['matricule'].id).last()
            response = {'voted': False}
            if vote:
                response['voted'] = True
            else:
                vote, created = Vote.objects.get_or_create(campaign=have_vote['campaign'],
                                                           elector=have_vote['matricule'], voted=False)
            serializer = VoteSerializers(vote, many=False)
            response['vote'] = serializer.data
            return Response(status=status.HTTP_200_OK, data=response)
        return Response(
            data=format_form_errors_response(form.errors),
            status=status.HTTP_400_BAD_REQUEST
        )

    @action(detail=False, methods=['post'], url_path='mark-vote')
    def mark_vote(self, request, *args, **kwargs):
        form = HaveVoteForm(request.data or {})
        if form.is_valid():
            have_vote = form.cleaned_data
            vote = Vote.objects.filter(voted=False, campaign__id=have_vote['campaign'].id,
                                       elector__id=have_vote['matricule'].id).last()
            if vote:
                vote.voted = True
                send_mt(elector=vote.elector)
                vote.save()
            else:
                vote = Vote.objects.create(campaign=have_vote['campaign'], elector=have_vote['matricule'], voted=False)
            serializer = VoteSerializers(vote, many=False)
            return Response(status=status.HTTP_200_OK, data=serializer.data)
        return Response(
            data=format_form_errors_response(form.errors),
            status=status.HTTP_400_BAD_REQUEST
        )

    @action(detail=False, methods=['get'])
    def export(self, request, *args, **kwargs):
        rows = self.filter_queryset(self.get_queryset()).values_list("elector__matricule", "elector__name",
                                                                     "elector__surname", "elector__birthDate",
                                                                     "elector__birthPlace", "elector__phone",
                                                                     "elector__gender", "elector__voteOffice__name",
                                                                     "elector__localisation", "campaign__name", "voted",
                                                                     "created__date", "created__time")

        rows1 = Elector.objects.exclude(id__in=[vote.elector.id for vote in Vote.objects.all()]).values_list(
            "matricule", "name", "surname", "birthDate",
            "birthPlace", "gender", "phone",
            "voteOffice__name", "localisation")
        rows1 = self.format_electors(rows=rows1)
        today_date = datetime.date.today()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="votes_' + str(today_date) + '.xlsx"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
        book = load_workbook(TEMPLATE_ROOT_VOTE_FILE)
        sheet = book['Votes']
        sheet, count = self.format_export_excel_rows(sheet=sheet, rows=rows)
        print(type(rows))
        sheet, count = self.format_export_excel_rows(sheet=sheet, rows=rows1, count=count)

        response.write(save_virtual_workbook(book))
        return response

    def format_electors(self, rows):
        response = []

        for row in rows:
            row = row + (None, "Absent", None, None)
            response.append(row)
        return response

    def format_export_excel_rows(self, sheet, rows, count=2):
        if rows:
            # Add rows
            for row in rows:
                start_letter = 65  # A
                for col_num in range(len(row)):
                    pos = chr(start_letter) + str(count)
                    value = row[col_num]
                    if isinstance(value, datetime.datetime):
                        value = value.strftime('%d/%m/%Y %H:%M:%S')
                        sheet[pos].style = datetime_style
                    elif isinstance(value, datetime.date):
                        value = value.strftime('%d/%m/%Y')
                        sheet[pos].style = date_style
                    elif isinstance(value, datetime.time):
                        value = value.strftime('%H:%M:%S')
                        sheet[pos].style = date_style
                    elif isinstance(value, bool):
                        value = 'Oui' if value else 'Non'
                    sheet[pos] = value
                    start_letter += 1
                count += 1
        return sheet, count


class ElectorView(viewsets.ModelViewSet):
    queryset = Elector.objects.none()
    serializer_class = ElectorSerializers
    pagination_class = CustomPagination
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = ElectorFilter
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        user = self.request.user
        if user.role == ROLE_SUPERVISOR:
            voteOffice = user.voteOffice
            if not voteOffice:
                return Elector.objects.none()
            return Elector.objects.filter(deleted=False, voteOffice__id=voteOffice.id)
        return Elector.objects.filter(deleted=False)

    @action(detail=False, methods=['post'], url_path='(?P<pk>.+)/picture')
    def upload_picture(self, request, *args, **kwargs):
        instance = self.get_object()
        if not request.FILES:
            form = UploadFileSerializer(data=request.data)
        else:
            form = UploadFileForm(request.data, request.FILES)
        if form.is_valid():
            directory = UPLOAD_DIR_ELECTOR + os.sep + instance.matricule
            if not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            handle_uploaded_file(
                request.FILES['picture'] if request.FILES else decode_base64_file(request.data['picture']),
                directory=directory, extra_date_folder=False)
            return Response(status=status.HTTP_200_OK)
        return Response(
            data=format_form_errors_response(form.errors),
            status=status.HTTP_400_BAD_REQUEST
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        directory = UPLOAD_DIR_ELECTOR + os.sep + instance.matricule
        if os.path.exists(directory):
            try:
                shutil.rmtree(directory)
            except:
                pass
        try:
            instance.delete()
        except:
            instance.deleted = True
            instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'])
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset()).values_list("matricule", "name", "surname", "birthDate",
                                                                         "birthPlace", "gender", "phone",
                                                                         "voteOffice__name", "localisation",
                                                                         "created__date", "created__time")

        today_date = datetime.date.today()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="electors_' + str(today_date) + '.xlsx"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition,X-Suggested-Filename'
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = self.__export_format(wb=wb, sheet_name='Electeurs',
                                  columns=["Matricule", "Nom", "Prenom", "Date naissance", "Lieu naissance", "Genre",
                                           "Téléphone", "Bureau vote", "Localisation", "Date", "Heure"],
                                  rows=queryset)
        wb.close()
        output.seek(0)
        response.write(output.read())
        return response

    def __export_format(self, wb, sheet_name, columns, rows):
        ws = wb.add_worksheet(sheet_name)
        row_num = 0
        font_style = wb.add_format({'bold': True})
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        font_style = wb.add_format({'bold': False})
        rows = [
            [x.strftime("%Y-%m-%d %H:%M") if isinstance(x, datetime.datetime) else x.strftime("%Y-%m-%d") if isinstance(
                x, datetime.datetime) or isinstance(x, datetime.date) else x for
             x in row] for row in
            rows]
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
        return ws

    @action(detail=False, methods=['get', 'delete'], url_path='(?P<pk>.+)/pictures')
    def get_pictures(self, request, *args, **kwargs):
        instance = self.get_object()
        matricule = instance.matricule
        if request.method == "GET":
            pictures = get_elector_pictures(matricule)
            responses = []
            if pictures:
                for picture in pictures:
                    if picture:
                        responses.append(MEDIA_URL + matricule + '/' + picture)
            return Response(status=status.HTTP_200_OK, data=responses)
        path = request.data.get('path', None)
        if path:
            paths = str(path).split(MEDIA_URL)
            if len(paths) > 1:
                directory = UPLOAD_DIR_ELECTOR + os.sep + paths[1].replace('/', os.sep)
                if os.path.exists(directory):
                    try:
                        os.remove(directory)
                    except:
                        pass
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['put'], url_path='(?P<pk>.+)/vote-office')
    def vote_office(self, request, *args, **kwargs):
        instance = self.get_object()
        form = AddVoteOffice(data=request.data or {})
        if form.is_valid():
            data = form.cleaned_data
            instance.voteOffice = data['voteOffice']
            instance.save()
            serializer = self.get_serializer(instance, many=False)
            return Response(status=status.HTTP_200_OK, data=serializer.data)
        return Response(
            data=format_form_errors_response(form.errors),
            status=status.HTTP_400_BAD_REQUEST
        )


class CampaignView(viewsets.ModelViewSet):
    queryset = Campaign.objects.filter(deleted=False)
    serializer_class = CampaignSerializers
    pagination_class = CustomPagination
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = CampaignFilter
    permission_classes = (IsAuthenticated,)


class VoteOfficeView(viewsets.ModelViewSet):
    queryset = VoteOffice.objects.filter(deleted=False)
    serializer_class = VoteOfficeSerializers
    pagination_class = CustomPagination
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = VoteOfficeFilter
    permission_classes = (IsAuthenticated,)
